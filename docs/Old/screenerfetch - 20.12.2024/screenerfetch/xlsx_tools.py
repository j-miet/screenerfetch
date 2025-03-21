import copy

import openpyxl
from openpyxl.styles import Font, Alignment

from query import COL_HEADERS
from file_paths import WB_PATH

def save(symbol_data, date_str):
    wb = openpyxl.load_workbook(WB_PATH)
    ws = wb.active
    next_row = ws.max_row+1
    full_data = copy.deepcopy(symbol_data)
    for symbol in full_data:
        symbol.insert(0, date_str)
    for row in full_data:
        ws.append(row)
        for i in range(1, len(row)+1):
            current = ws.cell(column=i, row=next_row)
            current.alignment = Alignment(horizontal='right')
        next_row += 1
    wb.save(WB_PATH)

def edit_notes(input_data):
    wb = openpyxl.load_workbook(WB_PATH)
    ws = wb.active
    date, symbol = input_data.split()[0], input_data.split()[1]
    for sym in ws['B']:
        if sym.value == symbol:
            if ws.cell(row=sym.row, column=1).value == date:
                notes = input('Your notes --> ')
                ws.cell(row=sym.row, column=len(COL_HEADERS), value=notes)
                print('Cell updated')
                wb.save(WB_PATH)
                return
    print('Failed to find cells corresponding to input data')

#add an image; 
# either manual save: place image to folder, call it by its name and save it to .xlsx, or
# use browser scraping tools, open browser, take screenshot, then save it .xlsx
def add_image():
    #TODO implement code below
    pass

def create_wb():
    wb = openpyxl.Workbook()
    wb.save(WB_PATH)
    wb = openpyxl.load_workbook(WB_PATH)
    ws_data = wb.active
    ws_data.title = 'Data'
    header_font = Font(name='Times New Roman', size=12, bold=True)
    for h in COL_HEADERS:
        ws_data[h] = COL_HEADERS[h]
        ws_data[h].font = header_font
        for i in range(1, len(COL_HEADERS)+1):
            current = ws_data.cell(column=i, row=1)
            current.alignment = Alignment(horizontal='right')
    wb.create_sheet('Analyze') 
    #TODO add a short text tutorial to 'Analyze' sheet on how to fetch data from 'Data' sheet
    #explain useful formulas/commands -> !Data, A1:B10, 'if'/'where'/comparison operators, XLOOKUP, VLOOKUP, MATCH etc.
    #also explain stuff like how to remove duplicate rows from 'Data' etc.
    wb.save(WB_PATH)