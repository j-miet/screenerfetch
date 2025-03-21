"""Workbook tools for current custom workbook."""

import datetime

import openpyxl
from openpyxl.styles import Alignment

from workbook_tools import WorkSheets, get_last_row
from paths import FilePaths

def add_row_in_sheet2(input_str: str) -> None:
    """Add a custom row for a symbol in the second worksheet.

    Requires a 'custom' workbook (or else you need to add columns manually).
     
    Customize the row/col values inside loop if you need to add/remove stuff.

    Args:
        input_str: Symbol name and date in SYMBOL YYYY-MM-DD format.
    """
    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkSheets.SHEET1]
    if ws is not None:
        symbol, date = input_str.split()
        raw_date = date.split('-')
        y, m, d = raw_date[0], raw_date[1], raw_date[2]
        date_datetime = datetime.datetime(int(y), int(m), int(d))
        for sym in ws['B']:
            if sym.value == symbol:
                if ws.cell(row=sym.row, column=1).value.date() == date_datetime.date(): # type: ignore              
                    ws = wb[WorkSheets.SHEET2]
                    next_row = get_last_row(WorkSheets.SHEET2)+1
                    ws.cell(column=1, row=next_row).value = date_datetime.date()
                    current = ws.cell(column=1, row=next_row)
                    current.alignment = Alignment(horizontal='left')
                    ws.cell(column=2, row=next_row).value = symbol
                    current = ws.cell(column=2, row=next_row)
                    current.alignment = Alignment(horizontal='right')

                    wb.save(FilePaths.wb_path)
                    print(f'Row for {symbol, date} added to sheet {WorkSheets.SHEET2}.')
                    return
    print('Failed to find cells corresponding to input data.')

def edit_notes(input_str: str) -> None:
    """Updates the second worksheet 'Notes' column values with input_data.
    
    Requires a 'custom' workbook (or else you need to add columns manually).
    
    datetime.datetime comparisons cause a lot of headache for type checker like mypy so type: ignore is added.

    Args:
        input_str: Symbol name and date in SYMBOL YYYY-MM-DD format.
    """
    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkSheets.SHEET2]
    if ws is not None:
        symbol, date = input_str.split()
        raw_date = date.split('-')
        y, m, d = raw_date[0], raw_date[1], raw_date[2]
        date_datetime = datetime.datetime(int(y), int(m), int(d))
        for sym in ws['B']:
            if sym.value == symbol:
                if ws.cell(row=sym.row, column=1).value.date() == date_datetime.date(): # type: ignore
                    notes = input('Your notes --> ')
                    ws.cell(row=sym.row, column=3, value=notes) # remember to update column if location changes.
                    ws.cell(row=sym.row, column=3, value=notes).alignment = Alignment(horizontal='right')
                    wb.save(FilePaths.wb_path)
                    print(f'Notes for {symbol, date} updated in sheet {WorkSheets.SHEET2}.')
                    return
    print('Failed to find cells corresponding to input data.')

def add_image_hyperlinks(input_str: str) -> None:
    """Add intraday and daily images for a symbol that has them available in custom/images folder.
    
    Requires a 'custom' workbook (or else you need to add columns manually).

    Images are added to the second worksheet. Will always add both image hyperlinks, even if one or both images are missing.

    Args:
        input_str: Symbol name and date in SYMBOL YYYY-MM-DD format.
    """
    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkSheets.SHEET2]
    if ws is not None:
        symbol, date = input_str.split()
        raw_date = date.split('-')
        y, m, d = raw_date[0], raw_date[1], raw_date[2]
        date_datetime = datetime.datetime(int(y), int(m), int(d))
        for sym in ws['B']:
            if sym.value == symbol:
                if ws.cell(row=sym.row, column=1).value.date() == date_datetime.date(): # type: ignore
                    r = str(sym.row)
                    ws['D'+r].hyperlink = FilePaths.PATH+'custom\\images\\'+symbol+' '+date+'.png'
                    ws['D'+r].value = 'Image'
                    ws['D'+r].style = "Hyperlink"
                    ws['D'+r].alignment = Alignment(horizontal='center')
                    ws['E'+r].hyperlink = FilePaths.PATH+'custom\\images\\'+symbol+' '+date+' D'+'.png'
                    ws['E'+r].value = 'Image'
                    ws['E'+r].style = "Hyperlink"
                    ws['E'+r].alignment = Alignment(horizontal='center')
                    wb.save(FilePaths.wb_path)
                    print('Done.')
                    return
    print('Failed to find cells corresponding to input data.')
    return