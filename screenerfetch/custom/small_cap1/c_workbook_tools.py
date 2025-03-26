"""Workbook tools for current custom workbook."""

import datetime
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Font, NamedStyle

from paths import FilePaths
from query import QueryVars
from custom.small_cap1.settings import SmallCap1Values
from sheets import WorkbookSheets
from workbook_tools import get_last_row

def _create_custom_workbook_files() -> None:
    try:
        for folder_path in (FilePaths.wb_files_path, FilePaths.data_path, FilePaths.settings_path):
            os.mkdir(folder_path)
        print(f"{FilePaths.wb_name}, {FilePaths.wb_name}/settings and {FilePaths.wb_name}/data created.")
    except FileExistsError:
        ...
    with open(FilePaths.settings_path/'settings.json', 'w') as f:
            set_settings = SmallCap1Values.SETTINGS
            json.dump(set_settings, f, indent=4)
    print(f"{FilePaths.wb_name}/settings/settings.json created.")
    with open(FilePaths.settings_path/'query.txt', 'w') as qf:
        qf.writelines(SmallCap1Values.QUERY)
    print(f"{FilePaths.wb_name}/settings/query.txt created")
    with open(FilePaths.settings_path/'headers.txt', 'w') as hf:
        hf.writelines(SmallCap1Values.HEADERS)
    print(f"{FilePaths.wb_name}/settings/headers.txt created.")

def add_row_in_sheet2(input_str: str) -> None:
    """Add a custom row for a symbol in the second worksheet.

    Requires a 'custom' workbook (or else you need to add columns manually).
     
    Customize the row/col values inside loop if you need to add/remove stuff.

    Args:
        input_str: Symbol name and date in SYMBOL YYYY-MM-DD format.
    """
    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkbookSheets.sheet_names[0]]
    if ws is not None:
        symbol, date = input_str.split()
        raw_date = date.split('-')
        y, m, d = raw_date[0], raw_date[1], raw_date[2]
        date_datetime = datetime.datetime(int(y), int(m), int(d))
        for sym in ws['B']:
            if sym.value == symbol:
                if ws.cell(row=sym.row, column=1).value.date() == date_datetime.date(): # type: ignore              
                    ws = wb[WorkbookSheets.sheet_names[1]]
                    next_row = get_last_row(WorkbookSheets.sheet_names[1])+1
                    ws.cell(column=1, row=next_row).value = date_datetime.date()
                    current = ws.cell(column=1, row=next_row)
                    current.alignment = Alignment(horizontal='left')
                    ws.cell(column=2, row=next_row).value = symbol
                    current = ws.cell(column=2, row=next_row)
                    current.alignment = Alignment(horizontal='right')

                    wb.save(FilePaths.wb_path)
                    print(f'Row for {symbol, date} added to sheet {WorkbookSheets.sheet_names[1]}.')
                    return
    print('Failed to find cells corresponding to input data.')

def edit_notes(input_str: str) -> None:
    """Updates the second worksheet 'Notes' column values with input_data.
    
    Requires a 'custom' workbook (or else you need to add columns manually).
    
    datetime.datetime comparisons cause a lot of headache for type checker like mypy so type: ignore is added.

    Args:
        input_str: Symbol name and date in SYMBOL YYYY-MM-DD format.
    """
    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkbookSheets.sheet_names[1]]
    if ws is not None:
        symbol, date = input_str.split()
        raw_date = date.split('-')
        y, m, d = raw_date[0], raw_date[1], raw_date[2]
        date_datetime = datetime.datetime(int(y), int(m), int(d))
        for sym in ws['B']:
            if sym.value == symbol:
                if ws.cell(row=sym.row, column=1).value.date() == date_datetime.date(): # type: ignore
                    notes = input('Your notes --> ')
                    ws.cell(row=sym.row, column=3, value=notes) # remember to update column if location changes.
                    ws.cell(row=sym.row, column=3, value=notes).alignment = Alignment(horizontal='right')
                    wb.save(FilePaths.wb_path)
                    print(f'Notes for {symbol, date} updated in sheet {WorkbookSheets.sheet_names[1]}.')
                    return
    print('Failed to find cells corresponding to input data.')

def add_image_hyperlinks(input_str: str) -> None:
    """Add intraday and daily images for a symbol that has them available in custom/images folder.
    
    Requires a 'custom' workbook (or else you need to add columns manually).

    Images are added to the second worksheet. Will always add both image hyperlinks, even if one or both images are missing.

    Args:
        input_str: Symbol name and date in SYMBOL YYYY-MM-DD format.
    """
    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkbookSheets.sheet_names[1]]
    if ws is not None:
        symbol, date = input_str.split()
        raw_date = date.split('-')
        y, m, d = raw_date[0], raw_date[1], raw_date[2]
        date_datetime = datetime.datetime(int(y), int(m), int(d))
        for sym in ws['B']:
            if sym.value == symbol:
                if ws.cell(row=sym.row, column=1).value.date() == date_datetime.date(): # type: ignore
                    r = str(sym.row)
                    ws['D'+r].hyperlink = FilePaths.PATH/'custom'/'images'/str(symbol+' '+date+'.png')
                    ws['D'+r].value = 'Image'
                    ws['D'+r].style = "Hyperlink"
                    ws['D'+r].alignment = Alignment(horizontal='center')
                    ws['E'+r].hyperlink = FilePaths.PATH/'custom'/'images'/str(symbol+' '+date+' D'+'.png')
                    ws['E'+r].value = 'Image'
                    ws['E'+r].style = "Hyperlink"
                    ws['E'+r].alignment = Alignment(horizontal='center')
                    wb.save(FilePaths.wb_path)
                    print('Done.')
                    return
    print('Failed to find cells corresponding to input data.')
    return

def custom_update_datetime() -> None:
    """Update datetime on second sheet."""
    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkbookSheets.sheet_names[1]]
    if ws is not None:
        for i in range(2, get_last_row(WorkbookSheets.sheet_names[1])+1):
            ws.cell(row=i, column=1).style = "datetime"
            ws.cell(row=i, column=1).alignment = Alignment(horizontal='left')          
            wb.save(FilePaths.wb_path)
            print('Date format updated.')

def create_custom_wb() -> None:
    """Create custom workbook. Will overwrite contents of currently existing wb with same name."""
    FilePaths.update_filepaths()
    _create_custom_workbook_files()
    QueryVars.update_query_variables()
    wb = openpyxl.Workbook()
    ws_data = wb.active
    if ws_data is not None:
        ws_data.title = "sheet1"
        header_font = Font(name='Times New Roman', size=12, bold=True)
        for h in QueryVars.col_headers:
            ws_data[h] = QueryVars.col_headers[h]
            ws_data[h].font = header_font
        for i in range(3, len(QueryVars.col_headers)+1):
            current = ws_data.cell(column=i, row=1)
            current.alignment = Alignment(horizontal='right')
        ws_data['A2'].style = NamedStyle(name="datetime", number_format='YYYY/MM/DD', 
                                            alignment=Alignment(horizontal='left'))
        ws_data.freeze_panes = 'A2'

        col_headers_sheet2= {
            'A1' : 'Date',
            'B1' : 'Symbol',
            'C1' : 'Notes',
            'D1' : 'Intraday',
            'E1' : 'Daily'
        }        
        wb.create_sheet("sheet2")
        ws_data = wb["sheet2"]
        for h in col_headers_sheet2:
            ws_data[h] = col_headers_sheet2[h]
            ws_data[h].font = header_font
        for i in range(2, len(col_headers_sheet2)+1):
            current = ws_data.cell(column=i, row=1)
            current.alignment = Alignment(horizontal='right')
        ws_data.freeze_panes = 'A2'
        wb.save(FilePaths.wb_path)

        WorkbookSheets.update_sheets()