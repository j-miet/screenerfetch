"""Functions for excel workbook data manipulation."""

import copy
import datetime
import json
import logging
import os

import openpyxl
from openpyxl.styles import Font, Alignment, NamedStyle
import pandas as pd

from paths import FilePaths
from query import QueryVars
from sheets import WorkbookSheets

logger = logging.getLogger('screenerfetch')

def _create_workbook_files() -> None:
    logger.debug("workbook_tools> _create_workbook_files")
    try:
        for folder_path in (FilePaths.wb_files_path, FilePaths.data_path, FilePaths.settings_path):
            os.mkdir(folder_path)
        print(f"{FilePaths.wb_name}, {FilePaths.wb_name}/settings and {FilePaths.wb_name}/data created.")
    except FileExistsError:
        logger.debug("workbook_tools> _create_workbook_files: At least one of the files already exists")
    with open(FilePaths.settings_path/'settings.json', 'w') as sf:
            set_settings = {"type": "basic", 
                            "market": "global",
                            "headers": {},
                            "query": {"columns": ["name"], "range": [0,1]}
            }
            json.dump(set_settings, sf, indent=4)
    print(f"{FilePaths.wb_name}/settings/settings.json created.")
    logger.debug("workbook_tools> _create_workbook_files: settings.json created")
    with open(FilePaths.settings_path/'query.txt', 'w') as qf:
        qf.writelines(('{\n', 
                       '    "columns": ["name"],\n', 
                       '    "range": [0,1]\n', 
                       '}'
                       ))
    print(f"{FilePaths.wb_name}/settings/query.txt created")
    logger.debug("workbook_tools> _create_workbook_files: query.txt created")
    with open(FilePaths.settings_path/'headers.txt', 'w') as hf:
        hf.writelines(('{\n', 
                       '   {}', 
                       '\n}'
                       ))
    print(f"{FilePaths.wb_name}/settings/headers.txt created.")
    logger.debug("workbook_tools> _create_workbook_files: headers.txt created")

def get_last_row(sheet_name: str) -> int:
    """Returns the last non-empty row index of main worksheet.

    Suprisingly, there's no automatic non-empty row counter function so I used the following because of its simplicity:
    https://singhaldhruv.medium.com/python-and-openpyxl-counting-non-empty-rows-in-excel-made-easy-36d708671918

    Args:
        sheet_name (str): Worksheet name.

    Returns:
        int:
        Last non-empty row.
    """
    logger.debug(f"workbook_tools> get_last_row: Sheet name '{sheet_name}'")
    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[sheet_name]
    rows = (len([row for row in ws if not all([row[0].value is None])]))
    wb.close()
    return rows

def check_date(date_str: str) -> bool:
    """Checks whether a date exists in workbook.
    
    Args:
        date_str (str): Date string in a yyyy-mm-dd format.
        
    Returns:
        bool:
        True if date found, else False.
    """
    logger.debug(f"workbook_tools> check_date: Date '{date_str}'")
    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkbookSheets.sheet_names[0]]
    for date in ws['A']:
        if date_str == str(date.value).replace(' 00:00:00', ''):
            return True
    print("Date value not found.")
    return False

def save(symbol_data: list[list[str]], date_str: str, auto_update_nums: bool = True) -> None:
    """Saves passed symbol_data to the main workbook file.
    
    Adds the date_str in front of symbol_data before adding all the symbol data workbook.

    Finally, perfoms an automatic str -> int/float conversion on selected columns; see more on that under update_values_to_nums(). You can disable this auto-update by setting AUTO_UPDATE_NUMS = False.

    Note: there's a faster way to add rows, with 'append' method, but it adds them after last visible row. 
    This means if you've scrolled the sheet down a lot, it will place the next row in a seemingly random
    row and leaves a gap of empty row in between.

    Args:
        symbol_data (list[list[str]]): 2D list where each inner list contains all data for a specific symbol.
            date_str: Date in DD/MM/YYYY format.
        date_str (str): Current date.
        auto_update_nums (bool): Whether to call auto-update current int and float columns. Default is True.
    """
    logger.debug("workbook_tools> save")

    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkbookSheets.sheet_names[0]]
    if ws is not None:
        d, m, y = date_str.split('/')
        starting_row = get_last_row(WorkbookSheets.sheet_names[0])+1
        next_row = starting_row
        full_data = copy.deepcopy(symbol_data)
        for row in full_data:
            for i, elem in zip(range(1, len(row)+1), row):
                ws.cell(column=i+1, row=next_row).value = elem
                ws.cell(column=i+1, row=next_row).alignment = Alignment(horizontal='right')
            current = ws.cell(column=1, row=next_row)
            current.value = datetime.datetime(int(y), int(m), int(d)).date()
            current.style = "datetime"
            current.alignment = Alignment(horizontal='left')
            next_row += 1
        wb.save(FilePaths.wb_path)
        logger.debug("workbook_tools> save: Data saved succesfully")
        if auto_update_nums:
            update_values_to_nums(starting_row)
        print('Saving succesful!')
        return
    print('Saving process failed.')

def update_values_to_nums(start_row: int = 2) -> None:
    """Update all values of listed columns to float/int type.
     
    Selected columns can be changed in query.py by editing CUSTOM_HEADERS.

    Remember to verify that 
     -column characters and numbers match AND
     -make a copy before running this first time, otherwise you could accidentally overwrite important data!

    Args:
        start_row (int=2): Row number where updating starts. Default is 2.
    """
    logger.debug(f"workbook_tools> update_values_to_nums: Start row {start_row}")
    int_columns = QueryVars.int_cols
    xlsx_int_columns = QueryVars.sheet_xlsx_int_cols
    float_columns = QueryVars.float_cols
    float_decimals = QueryVars.float_decimals
    xlsx_float_columns = QueryVars.sheet_xlsx_float_cols

    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkbookSheets.sheet_names[0]]
    if ws is not None:
        for r in range(start_row, get_last_row(WorkbookSheets.sheet_names[0])+1):
            for f_col in range(len(float_columns)):
                try:
                    decimals = float_decimals[float_columns[f_col]]
                    ws[float_columns[f_col][0]+str(r)] = (
                        f"{float((ws.cell(row=r, column=xlsx_float_columns[f_col]).value)):.{decimals}f}") # type: ignore
                except (TypeError, ValueError, KeyError):
                    ...
            for i_col in range(len(int_columns)):
                try:
                    ws[int_columns[i_col][0]+str(r)].number_format = '0'
                    ws[int_columns[i_col][0]+str(r)] = (
                        int(ws.cell(row=r, column=xlsx_int_columns[i_col]).value)) # type: ignore
                except (TypeError, ValueError):
                    ...
        wb.save(FilePaths.wb_path)
        logger.debug("workbook_tools> update_values_to_nums: Numerical values updated.")
        print('Values updated to numbers.')
        return

def update_datetime(first_row: int = 2) -> None:
    """Update date column values to datetime yyyy-mm-dd format.
    
    Sometimes cells update their values to correct yyyy-mm-dd format, and get the full
    yyyy-mm-dd;hh-mm-ss format. To remove the non-date part, this function will go through all
    the column values, starting from first_row and ending on last non-empty row value.

    first_row gets passed a base value 2 (1 would be the date header), but this value can be
    any value >= 2.

    Currently updates *all* dates in the second sheet as well.
    
    Args:
        first_row (int=2): Row number where updating begins. Is always >= 2 as row 1 points to headers.
    """
    logger.debug(f"workbook_tools> update_datetime: First row {first_row}")
    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkbookSheets.sheet_names[0]]
    if ws is not None:
        if isinstance(first_row, int) and first_row >= 2:
            for i in range(first_row, get_last_row(WorkbookSheets.sheet_names[0])+1):
                ws.cell(row=i, column=1).style = "datetime"
                ws.cell(row=i, column=1).alignment = Alignment(horizontal='left')            
            wb.save(FilePaths.wb_path)
            print('Date format updated.')
            return
    print('Argument must be an integer value greater than or equal to 2 (>= 2).')

def update_headers() -> None:
    """Update xlsx workbook file header columns."""
    logger.debug("workbook_tools> update_headers: Updating workbook headers")
    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkbookSheets.sheet_names[0]]
    if ws is not None:
        header_font = Font(name='Times New Roman', size=12, bold=True)
        for h in QueryVars.col_headers.keys():
            ws[h] = QueryVars.col_headers[h]
            ws[h].font = header_font
        for i in range(2, len(QueryVars.col_headers.keys())+1):
            current = ws.cell(column=i, row=1)
            current.alignment = Alignment(horizontal='right')
        wb.save(FilePaths.wb_path)
        print("Xlsx file headers updated.")
    return

def export_wb(type: str) -> None:
    """Exports workbook data to specific format'.
    
    Args:
        type (str): File format - 'txt', 'csv' or 'json'. Can also pass 'all' to export all supported file types.
    """
    logger.debug(f"workbook_tools> export_wb: Output file type '{type}'")
    df = pd.read_excel(FilePaths.wb_path, WorkbookSheets.sheet_names[0])
    if type == 'txt': 
        df.to_csv(FilePaths.data_path/str(FilePaths.wb_name+'.txt'), sep='\t', index=False)
    elif type == 'csv':
        df.to_csv(FilePaths.data_path/str(FilePaths.wb_name+'.csv'), index=False)
    elif type == 'json':
        df.to_json(FilePaths.data_path/str(FilePaths.wb_name+'.json'), indent=1)
    elif type == 'all':
        df.to_csv(FilePaths.data_path/str(FilePaths.wb_name+'.txt'), sep='\t', index=False)
        df.to_csv(FilePaths.data_path/str(FilePaths.wb_name+'.csv'), index=False)
        df.to_json(FilePaths.data_path/str(FilePaths.wb_name+'.json'), indent=1)
        print(f"Created txt, csv and json files in {FilePaths.wb_name}/data.")
        return
    else:
        print("Invalid file type.")
        return
    print(f'{FilePaths.wb_name}.'+type+f' created in {FilePaths.wb_name}/data folder.')

def remove_duplicates() -> None:
    """Remove duplicate rows from current .xlsx file.
    
    Uses date and symbol name to differentiate rows: one symbol cannot exists twice on same day.
    """
    logger.debug("workbook_tools> remove_duplicates")
    wb = openpyxl.load_workbook(FilePaths.wb_path)
    ws = wb[WorkbookSheets.sheet_names[0]]
    counter = 0
    if ws is not None:
        for row in reversed(list(ws.iter_rows(min_row=2, min_col=1, max_col=2))):
            current = (row[0].value, row[1].value)
            for row2 in ws.iter_rows(min_row=2, min_col=1, max_col=2):
                match = (row2[0].value, row2[1].value)
                if (row[0].row is not None 
                    and row[0].row != row2[0].row
                    and current[0] is not None
                    and current[0] == match[0] 
                    and current[1] == match[1]):
                    ws.delete_rows(row[0].row, 1)
                    counter += 1
                    print(f"Removed row number {row[0].row}.")
                    break
    logger.debug(f"workbook_tools> update_values_to_nums: Duplicates removed: {counter}")
    if counter == 1:
        print("1 row was removed.")
    else:
        print(f"{counter} rows were removed.")
    if counter > 0:
        wb.save(FilePaths.wb_path)
                
def create_wb(new_files: bool = True) -> None:
    """Creates a new workbook main file and names the worksheet with sheets.WorkbookSheetNames.sheet_names[0] value.
    
    Calling this function will overwrite the contents of existing workbook file.

    Args:
        new_files (bool=True): Boolean to determine whether old settings files are overwritten or not; if False, will 
            keep files which enabled custom columns for workbook.
    """
    if new_files:
        logger.debug("workbook_tools> create_wb: File overwrite enabled")
    else:
        logger.debug("workbook_tools> create_wb")
    FilePaths.update_filepaths()
    if new_files:
        _create_workbook_files()
    QueryVars.update_query_variables()
    wb = openpyxl.Workbook()
    ws_data = wb.active
    if ws_data is not None:
        ws_data.title = "sheet1"
        header_font = Font(name='Times New Roman', size=12, bold=True)
        for h in QueryVars.col_headers.keys():
            ws_data[h] = QueryVars.col_headers[h]
            ws_data[h].font = header_font
        for i in range(2, len(QueryVars.col_headers.keys())+1):
            current = ws_data.cell(column=i, row=1)
            current.alignment = Alignment(horizontal='right')
        # before a style can be used, it has to be allocated to any cell. After this, any future cells can simply 
        # copy the style by using self.style = "datetime".
        ws_data['A2'].style = NamedStyle(name="datetime", number_format='YYYY/MM/DD')
        ws_data.freeze_panes = 'A2'
        wb.save(FilePaths.wb_path)
        if new_files:
            print(f"New workbook {FilePaths.wb_name}.xlsx created.")
        WorkbookSheets.update_sheets()