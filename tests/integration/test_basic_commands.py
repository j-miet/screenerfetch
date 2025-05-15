"""Testing some of the basic screenerfetch commands.

Test proceeds as follows (commands in parenthesis)  
-create a new workbook (wb)  
-set query settings, update variables (q, UPDATE WB)  
-fetch screener data from web api (f)  
-save data to xlsx file (saveall)  
-test duplicate data deletion, then delete workbook. (remove duplicates, DELETE WB)

Creating and deleting are tested more thoroughly under test_create_and_delete.py.
"""

import datetime
import json
import os

import openpyxl
import pytest

import commands
import commands_utils
from paths import FilePaths
from query import QueryVars, FetchData
import workbook_tools

class TestVals:
    """Stores test wb name and previous wb name prior to running tests."""
    WB_NAME_STR = 'integration_test_wb'
    previous_wb: dict[str, str]


def test_create_workbook():
    wb_name: dict[str, str]
    # if previous workbook folder exists after failed test, delete it
    commands_utils.delete_workbook(TestVals.WB_NAME_STR)

    # workbook creation
    with open(FilePaths.WB_FILES_ROOT_PATH/'current_wb.json') as wb_name_file:
        wb_name = json.load(wb_name_file)
    TestVals.previous_wb = wb_name.copy()
    assert wb_name["wb_name"] == TestVals.previous_wb["wb_name"]

    assert commands_utils.change_workbook(TestVals.WB_NAME_STR, False) == -2
    assert commands_utils.change_workbook(TestVals.WB_NAME_STR, True) == 2
    assert commands_utils.change_workbook(TestVals.WB_NAME_STR, False) == 1
    commands_utils.change_workbook(TestVals.WB_NAME_STR, True) == -3

    assert FilePaths.wb_name == TestVals.WB_NAME_STR
    assert FilePaths.wb_path == FilePaths.WB_FILES_ROOT_PATH/TestVals.WB_NAME_STR/str(TestVals.WB_NAME_STR+'.xlsx')
    assert FilePaths.wb_name in os.listdir(FilePaths.WB_FILES_ROOT_PATH)
    assert FilePaths.settings_path == FilePaths.WB_FILES_ROOT_PATH/TestVals.WB_NAME_STR/'settings'
    assert {'query.txt', 'headers.txt', 'settings.json'}.issubset(os.listdir(FilePaths.settings_path))
    assert FilePaths.data_path == FilePaths.WB_FILES_ROOT_PATH/TestVals.WB_NAME_STR/'data'

    assert QueryVars.my_query == {"columns": ["name"], "range": [0,1]}
    assert QueryVars.market == 'global'
    assert QueryVars.custom_headers == {}

    # query data
    new_query = {"columns": ["name", "open"], "range": [0,10]}
    with open(FilePaths.settings_path/'query.txt', 'w') as wb_query:
        json.dump(new_query, wb_query, indent=4)
    commands_utils.update_settings_json('query', False)

    new_headers = {"A": {"name": "Date Value"},
                    "B": {"name": "Symbol Name"},
                    "C": {"name": "Open Price", "type": "float"}}
    with open(FilePaths.settings_path/'headers.txt', 'w') as wb_headers:
        json.dump(new_headers, wb_headers, indent=4)
    commands_utils.update_settings_json('headers', False)

    with open(FilePaths.settings_path/'settings.json') as wb_settings:
        settings = json.load(wb_settings)
    settings["market"] = 'germany'
    with open(FilePaths.settings_path/'settings.json', 'w') as wb_settings:
        json.dump(settings, wb_settings, indent=4)
    QueryVars.update_query_variables()

    with open(FilePaths.settings_path/'settings.json') as wb_settings:
        settings = json.load(wb_settings)
    assert settings["type"] == 'basic'
    assert settings["market"] == 'germany'
    assert settings["headers"] == new_headers
    assert settings["query"] == new_query

    # fetch web data
    assert FilePaths.wb_name == TestVals.WB_NAME_STR 
    assert FilePaths.wb_path == FilePaths.WB_FILES_ROOT_PATH/TestVals.WB_NAME_STR/str(TestVals.WB_NAME_STR+'.xlsx')

    with open(FilePaths.settings_path/'settings.json') as wb_settings:
        settings = json.load(wb_settings)
    settings["market"] = '-'
    with open(FilePaths.settings_path/'settings.json', 'w') as wb_settings:
        json.dump(settings, wb_settings, indent=4)
    QueryVars.update_query_variables()
    with pytest.raises(Exception, match=r"Could not fetch any data from API."): # first, test with bad market value
        commands_utils.requests_api_data()

    with open(FilePaths.settings_path/'settings.json') as wb_settings: # insert valid market value back
        settings = json.load(wb_settings)
    settings["market"] = 'germany'
    with open(FilePaths.settings_path/'settings.json', 'w') as wb_settings:
        json.dump(settings, wb_settings, indent=4)
    QueryVars.update_query_variables()
    assert QueryVars.market == 'germany'
    assert commands.fetch() == 0 # then make a post request to actual Tradingview public web api

    assert len(FetchData.query_data) > 0
    with open(FilePaths.TXT_PATH) as api_data_txt:
        api_data = api_data_txt.readlines()
    assert len(api_data) == 6 + len(FetchData.query_data)
    for symbol in FetchData.query_data:
        assert len(symbol) == 2

    # xslx file: saving data
    # current workbook hasn't been updated yet for custom headers
    test_wb = openpyxl.load_workbook(FilePaths.wb_path)
    assert len(test_wb.sheetnames) == 1
    assert test_wb.sheetnames[0] == 'sheet1'
    test_wb_sheet = test_wb["sheet1"]
    assert test_wb_sheet["A1"].value == 'date'
    assert test_wb_sheet["B1"].value == 'name'
    assert test_wb_sheet["C1"].value == 'open'
    assert test_wb_sheet["A2"].value == None
    assert test_wb_sheet["B2"].value == None
    commands.saveall()  # save data
    test_wb = openpyxl.load_workbook(FilePaths.wb_path)
    test_wb_sheet = test_wb["sheet1"]
    assert len(test_wb_sheet["A"]) == 1 + len(FetchData.query_data) # column + all data rows
    assert len(test_wb_sheet["B"]) == 1 + len(FetchData.query_data)
    assert test_wb_sheet["C1"].value == 'open'
    assert len(test_wb_sheet["C"]) == 1 + len(FetchData.query_data) # data still stored under column C

    # format workbook data, but don't delete settings files; simulates 'UPDATE WB' commands.
    commands.create(False)
    test_wb = openpyxl.load_workbook(FilePaths.wb_path)
    test_wb_sheet = test_wb["sheet1"]
    assert test_wb_sheet["A1"].value == 'Date Value'
    assert test_wb_sheet["B1"].value == 'Symbol Name'
    assert test_wb_sheet["C1"].value == 'Open Price'
    assert test_wb_sheet["A2"].value == None
    assert test_wb_sheet["B2"].value == None
    assert test_wb_sheet["C2"].value == None
    commands.saveall()
    test_wb = openpyxl.load_workbook(FilePaths.wb_path)
    test_wb_sheet = test_wb["sheet1"]
    
    y, m, d = str(datetime.date.today()).split('-')
    assert test_wb_sheet["A2"].value == datetime.datetime(int(y),int(m), int(d), 0, 0)
    assert test_wb_sheet["B2"].value == FetchData.query_data[0][0]
    assert test_wb_sheet["C2"].value == FetchData.query_data[0][1]
    assert len(test_wb_sheet["A"]) == 1 + len(FetchData.query_data)
    assert len(test_wb_sheet["B"]) == 1 + len(FetchData.query_data)
    assert len(test_wb_sheet["C"]) == 1 + len(FetchData.query_data)
    if FetchData.query_data == 10:
        assert test_wb_sheet["A11"].value == datetime.datetime(int(y),int(m), int(d), 0, 0)
        assert test_wb_sheet["B11"].value == FetchData.query_data[9][0]
        assert test_wb_sheet["C11"].value == FetchData.query_data[9][1]

    commands.saveall()
    test_wb = openpyxl.load_workbook(FilePaths.wb_path)
    test_wb_sheet = test_wb["sheet1"]
    assert test_wb_sheet["A2"].value != None
    assert test_wb_sheet["A2"].value  == test_wb_sheet["A12"].value
    if FetchData.query_data == 10:
        assert test_wb_sheet["A11"].value != None
        assert test_wb_sheet["A11"].value  == test_wb_sheet["A21"].value

    # test xslx commands, then delete wb
    workbook_tools.export_wb('all')
    assert {TestVals.WB_NAME_STR+'.txt', 
            TestVals.WB_NAME_STR+'.csv',
            TestVals.WB_NAME_STR+'.json'}.issubset(os.listdir(FilePaths.data_path))

    test_wb_sheet = openpyxl.load_workbook(FilePaths.wb_path)["sheet1"]
    total_rows = len(test_wb_sheet["A"])
    commands.remove_duplicate_data()
    test_wb_sheet = openpyxl.load_workbook(FilePaths.wb_path)["sheet1"]
    assert len(test_wb_sheet["A"]) == total_rows-len(FetchData.query_data)
    
    commands_utils.delete_workbook(TestVals.WB_NAME_STR)

    # switch to workbook that was used before testing
    with open(FilePaths.WB_FILES_ROOT_PATH/'current_wb.json', 'w') as wb_name_file:
        json.dump(TestVals.previous_wb, wb_name_file, indent=4)